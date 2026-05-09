import streamlit as st
import pandas as pd
import plotly.express as px
import io
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font

st.set_page_config(page_title="Dashboard Producción — MetalParts", layout="wide")

@st.cache_data
def cargar_datos():
    df = pd.read_csv("caso3_produccion_dataset.csv")
    df["fecha_produccion"] = pd.to_datetime(df["fecha_produccion"])
    return df

df = cargar_datos()

# ── Sidebar filtros ──────────────────────────────────────────
st.sidebar.header("Filtros")

lineas = st.sidebar.multiselect(
    "Línea de producción",
    options=sorted(df["linea_produccion"].unique()),
    default=sorted(df["linea_produccion"].unique()),
)

turnos = st.sidebar.multiselect(
    "Turno",
    options=["Mañana", "Tarde", "Noche"],
    default=["Mañana", "Tarde", "Noche"],
)

fecha_min = df["fecha_produccion"].min().date()
fecha_max = df["fecha_produccion"].max().date()
rango_fecha = st.sidebar.date_input(
    "Rango de fechas",
    value=(fecha_min, fecha_max),
    min_value=fecha_min,
    max_value=fecha_max,
)

# Aplicar filtros
fecha_inicio = pd.Timestamp(rango_fecha[0])
fecha_fin = pd.Timestamp(rango_fecha[1] if len(rango_fecha) == 2 else rango_fecha[0])
dff = df[
    df["linea_produccion"].isin(lineas)
    & df["turno"].isin(turnos)
    & df["fecha_produccion"].between(fecha_inicio, fecha_fin)
]

# ── Título ───────────────────────────────────────────────────
st.title("🏭 Dashboard de Control de Producción — MetalParts Colombia")
st.caption(f"Mostrando {len(dff):,} órdenes de {len(df):,} totales")

# ── Tabs ─────────────────────────────────────────────────────
tab1, tab2, tab3 = st.tabs(["📊 Resumen", "📈 Tendencias", "⚠️ Alertas"])

with tab1:
    # ── KPIs (PASO 3) ────────────────────────────────────────
    eficiencia_promedio    = dff["eficiencia_pct"].mean()
    tasa_defectos_promedio = dff["tasa_defectos_pct"].mean()
    total_producidas       = dff["unidades_producidas"].sum()
    costo_total            = dff["costo_produccion_cop"].sum()
    tiempo_paro_total      = dff["tiempo_paro_min"].sum()
    horas_paro             = tiempo_paro_total / 60

    c1, c2, c3, c4, c5 = st.columns(5)
    c1.metric("Eficiencia promedio",       f"{eficiencia_promedio:.2f}%")
    c2.metric("Tasa de defectos promedio", f"{tasa_defectos_promedio:.2f}%")
    c3.metric("Total producido",           f"{total_producidas:,.0f}")
    c4.metric("Costo producción",          f"${costo_total:,.0f} COP")
    c5.metric("Tiempo de paro",            f"{tiempo_paro_total:,.1f} min ({horas_paro:.1f} h)")

    st.divider()

    # ── Agrupaciones (PASO 4) ────────────────────────────────
    efic_linea = (
        dff.groupby("linea_produccion")["eficiencia_pct"]
        .mean()
        .round(2)
        .sort_values(ascending=False)
        .reset_index()
        .rename(columns={"eficiencia_pct": "eficiencia_promedio"})
    )

    defectos_turno = (
        dff.groupby("turno")["unidades_defectuosas"]
        .sum()
        .sort_values(ascending=False)
        .reset_index()
    )

    paro_maquina = (
        dff.groupby("maquina")["tiempo_paro_min"]
        .sum()
        .round(1)
        .sort_values(ascending=False)
        .reset_index()
        .rename(columns={"tiempo_paro_min": "tiempo_paro_total_min"})
    )
    paro_maquina["tiempo_paro_horas"] = (paro_maquina["tiempo_paro_total_min"] / 60).round(2)

    col_a, col_b = st.columns(2)

    with col_a:
        # Fig1 — PASO 5: barras horizontales eficiencia por línea
        fig1 = px.bar(
            efic_linea.sort_values("eficiencia_promedio", ascending=True),
            y                      = "linea_produccion",
            x                      = "eficiencia_promedio",
            orientation            = "h",
            title                  = "🏭 Eficiencia Promedio por Línea de Producción (%)",
            labels                 = {"eficiencia_promedio": "Eficiencia Promedio (%)", "linea_produccion": ""},
            color                  = "eficiencia_promedio",
            color_continuous_scale = "Blues",
            text_auto              = ".2s",
        )
        fig1.update_layout(height=350, showlegend=False)
        st.plotly_chart(fig1, use_container_width=True)

    with col_b:
        # Fig2 — PASO 6: barras de defectos por turno
        fig2 = px.bar(
            defectos_turno.sort_values("unidades_defectuosas", ascending=False),
            x                       = "turno",
            y                       = "unidades_defectuosas",
            color                   = "turno",
            title                   = "⚠️ Total de Unidades Defectuosas por Turno",
            labels                  = {"unidades_defectuosas": "Total de Defectos", "turno": "Turno"},
            text_auto               = ".2s",
            color_discrete_sequence = px.colors.qualitative.Pastel,
        )
        fig2.update_layout(height=450, showlegend=False)
        st.plotly_chart(fig2, use_container_width=True)

    col_c, col_d = st.columns(2)

    with col_c:
        # Fig3 — PASO 7: barras horizontales paro por máquina
        fig3 = px.bar(
            paro_maquina.sort_values("tiempo_paro_total_min", ascending=True),
            y                      = "maquina",
            x                      = "tiempo_paro_total_min",
            orientation            = "h",
            title                  = "⏱️ Tiempo de Paro Total por Máquina (min)",
            labels                 = {"tiempo_paro_total_min": "Tiempo de Paro (min)", "maquina": ""},
            color                  = "tiempo_paro_total_min",
            color_continuous_scale = "Reds",
            text_auto              = ".2s",
        )
        fig3.update_layout(height=350, showlegend=False)
        st.plotly_chart(fig3, use_container_width=True)

    with col_d:
        # Fig4 — PASO 8: línea de producción semanal
        produccion_semanal_resumen = (
            dff.groupby("semana")
            .agg(unidades_producidas=("unidades_producidas", "sum"))
            .reset_index()
            .sort_values("semana")
        )
        fig4 = px.line(
            produccion_semanal_resumen,
            x       = "semana",
            y       = "unidades_producidas",
            title   = "📈 Evolución de la Producción Semanal",
            labels  = {"semana": "Semana del Año", "unidades_producidas": "Total de Unidades Producidas"},
            markers = True,
        )
        fig4.update_traces(
            line_color = "#2563eb",
            line_width = 3,
            marker     = dict(size=8, color="#1e40af"),
        )
        fig4.update_layout(
            height        = 350,
            xaxis         = dict(tickmode="linear", dtick=4),
            yaxis         = dict(gridcolor="#e2e8f0"),
            plot_bgcolor  = "#f8fafc",
            paper_bgcolor = "white",
        )
        st.plotly_chart(fig4, use_container_width=True)

    # Fig5 — PASO 9: scatter temperatura vs defectos con trendline (ancho completo)
    fig5 = px.scatter(
        dff,
        x          = "temperatura_c",
        y          = "tasa_defectos_pct",
        color      = "linea_produccion",
        trendline  = "ols",
        title      = "🌡️ Relación entre Temperatura y Tasa de Defectos",
        labels     = {
            "temperatura_c":     "Temperatura (°C)",
            "tasa_defectos_pct": "Tasa de Defectos (%)",
            "linea_produccion":  "Línea",
        },
        hover_data = ["id_orden", "producto", "turno"],
    )
    fig5.update_layout(
        height        = 450,
        plot_bgcolor  = "#f8fafc",
        paper_bgcolor = "white",
        xaxis         = dict(gridcolor="#e2e8f0"),
        yaxis         = dict(gridcolor="#e2e8f0"),
        legend        = dict(title="Línea", bgcolor="white", bordercolor="#e2e8f0", borderwidth=1),
    )
    st.plotly_chart(fig5, use_container_width=True)

with tab2:
    # Producción semanal (PASO 4 + PASO 8)
    produccion_semanal = (
        dff.groupby("semana")
        .agg(
            unidades_producidas = ("unidades_producidas", "sum"),
            ordenes             = ("id_orden", "count"),
            eficiencia_promedio = ("eficiencia_pct", "mean"),
        )
        .round(2)
        .reset_index()
        .sort_values("semana")
    )
    produccion_semanal["eficiencia_promedio"] = produccion_semanal["eficiencia_promedio"].round(2)

    # Fig4 — PASO 8: línea de producción semanal
    fig4 = px.line(
        produccion_semanal,
        x       = "semana",
        y       = "unidades_producidas",
        title   = "📈 Evolución de la Producción Semanal",
        labels  = {"semana": "Semana del Año", "unidades_producidas": "Total de Unidades Producidas"},
        markers = True,
    )
    fig4.update_traces(
        line_color = "#2563eb",
        line_width = 3,
        marker     = dict(size=8, color="#1e40af"),
    )
    fig4.update_layout(
        height        = 400,
        xaxis         = dict(tickmode="linear", dtick=4),
        yaxis         = dict(gridcolor="#e2e8f0"),
        plot_bgcolor  = "#f8fafc",
        paper_bgcolor = "white",
    )
    st.plotly_chart(fig4, use_container_width=True)

    fig6 = px.line(
        produccion_semanal,
        x       = "semana",
        y       = "eficiencia_promedio",
        title   = "📊 Eficiencia Promedio Semanal (%)",
        labels  = {"semana": "Semana del Año", "eficiencia_promedio": "Eficiencia Promedio (%)"},
        markers = True,
    )
    fig6.update_traces(line_color="#16a34a", line_width=3, marker=dict(size=8, color="#15803d"))
    fig6.update_layout(
        height        = 400,
        xaxis         = dict(tickmode="linear", dtick=4),
        yaxis         = dict(gridcolor="#e2e8f0"),
        plot_bgcolor  = "#f8fafc",
        paper_bgcolor = "white",
    )
    st.plotly_chart(fig6, use_container_width=True)

with tab3:
    alertas = dff[dff["tasa_defectos_pct"] > 10][
        ["id_orden", "fecha_produccion", "linea_produccion", "turno",
         "maquina", "producto", "tasa_defectos_pct", "unidades_defectuosas"]
    ].sort_values("tasa_defectos_pct", ascending=False)

    st.subheader(f"Órdenes con tasa de defectos > 10%  ({len(alertas)} registros)")
    st.dataframe(alertas, use_container_width=True)

    def to_excel(df: pd.DataFrame) -> bytes:
        buf = io.BytesIO()
        with pd.ExcelWriter(buf, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name="Producción")
            ws = writer.sheets["Producción"]
            # Formato tabla
            from openpyxl.worksheet.table import Table, TableStyleInfo
            ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
            tbl = Table(displayName="TablaProduccion", ref=ref)
            tbl.tableStyleInfo = TableStyleInfo(
                name="TableStyleMedium9", showFirstColumn=False,
                showLastColumn=False, showRowStripes=True, showColumnStripes=False,
            )
            ws.add_table(tbl)
            # Cabecera en negrita y color
            header_fill = PatternFill("solid", fgColor="2563EB")
            header_font = Font(bold=True, color="FFFFFF")
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
            # Ajustar ancho de columnas
            for col in ws.columns:
                max_len = max(len(str(c.value)) if c.value else 0 for c in col)
                ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 40)
        return buf.getvalue()

    st.download_button(
        "⬇️ Descargar dataset filtrado (.xlsx)",
        data      = to_excel(dff),
        file_name = "produccion_filtrado.xlsx",
        mime      = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
